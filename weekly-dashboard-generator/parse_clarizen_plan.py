from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


REQUIRED_HEADERS = {"State", "Name", "Start Date", "Level"}
END_DATE_HEADERS = ("Due Date", "End Date")
ADMIN_TASK_PHRASES = (
    "project management",
    "weekly project meetings",
    "status meeting",
    "pmo",
    "admin",
    "internal meeting",
    "update devops",
    "post go-live support",
)
INVOICING_MILESTONE_RE = re.compile(r"^(?:MS|M)-\d+", re.IGNORECASE)


class ClarizenPlanError(ValueError):
    """Raised when a Clarizen workbook cannot be parsed."""


def clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def find_header_row(rows) -> tuple[int, dict[str, int], str]:
    for row_number, row in enumerate(rows, start=1):
        headers = [clean(value) for value in row]
        header_map = {
            header: index for index, header in enumerate(headers) if header
        }
        if not REQUIRED_HEADERS.issubset(header_map):
            continue
        end_header = next(
            (header for header in END_DATE_HEADERS if header in header_map), ""
        )
        if not end_header:
            raise ClarizenPlanError(
                "Found the Clarizen header row, but could not find a Due Date "
                "or End Date column."
            )
        return row_number, header_map, end_header
    raise ClarizenPlanError(
        "Could not find the Clarizen header row containing State, Name, "
        "Start Date, Due Date/End Date, and Level."
    )


def parse_level(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if not number.is_integer():
        return None
    return int(number)


def parse_excel_date(value: Any, epoch) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            converted = from_excel(value, epoch)
            return converted.date() if isinstance(converted, datetime) else converted
        except (TypeError, ValueError, OverflowError):
            return None
    text = clean(value)
    for date_format in (
        "%Y-%m-%d",
        "%m/%d/%Y",
        "%m/%d/%y",
        "%B %d, %Y",
        "%b %d, %Y",
    ):
        try:
            return datetime.strptime(text, date_format).date()
        except ValueError:
            continue
    return None


def rank_deliverables(
    deliverables: list[dict[str, Any]],
    reference_date: date,
    recommendation_limit: int = 5,
) -> list[dict[str, Any]]:
    ranked = []
    window_end = reference_date + timedelta(days=60)

    for item in deliverables:
        start_date = (
            date.fromisoformat(item["start_date"]) if item.get("start_date") else None
        )
        end_date = (
            date.fromisoformat(item["end_date"]) if item.get("end_date") else None
        )
        name = item.get("name", "")
        is_admin = any(
            phrase in name.casefold() for phrase in ADMIN_TASK_PHRASES
        )

        if is_admin:
            rank_group = 5
            ranking_reason = "Long-running/admin task"
        elif start_date and end_date and start_date <= reference_date <= end_date:
            rank_group = 0
            ranking_reason = "Currently active"
        elif end_date and reference_date < end_date <= window_end:
            rank_group = 1
            ranking_reason = "Ending within 60 days"
        elif start_date and reference_date < start_date <= window_end:
            rank_group = 2
            ranking_reason = "Starting within 60 days"
        elif not start_date or not end_date:
            rank_group = 3
            ranking_reason = "Missing dates"
        else:
            rank_group = 4
            ranking_reason = "Older active task"

        ranked_item = dict(item)
        ranked_item["ranking_reason"] = ranking_reason
        ranked_item["_rank_group"] = rank_group
        ranked.append(ranked_item)

    ranked.sort(
        key=lambda item: (
            item["_rank_group"],
            not bool(item.get("end_date")),
            item.get("end_date") or "9999-12-31",
            not bool(item.get("start_date")),
            item.get("start_date") or "9999-12-31",
            item.get("name", "").casefold(),
        )
    )
    for index, item in enumerate(ranked):
        item["recommended"] = index < recommendation_limit
        item.pop("_rank_group", None)
    return ranked


def parse_clarizen_workbook(
    workbook_bytes: bytes,
    today: date | None = None,
    reference_date: date | None = None,
    file_name: str = "",
) -> tuple[list[dict[str, Any]], list[str]]:
    today = today or date.today()
    reference_date = reference_date or today
    is_legacy_xls = file_name.lower().endswith(".xls")
    if is_legacy_xls:
        try:
            import xlrd

            workbook = xlrd.open_workbook(file_contents=workbook_bytes)
            worksheet = (
                workbook.sheet_by_name("Work Plan")
                if "Work Plan" in workbook.sheet_names()
                else workbook.sheet_by_index(0)
            )
            rows = [worksheet.row_values(index) for index in range(worksheet.nrows)]

            def convert_date(value):
                if value in (None, ""):
                    return None
                if isinstance(value, (int, float)):
                    try:
                        return xlrd.xldate_as_datetime(
                            value, workbook.datemode
                        ).date()
                    except (ValueError, OverflowError):
                        return None
                return parse_excel_date(value, None)

        except ImportError as exc:
            raise ClarizenPlanError(
                "Legacy .xls support requires xlrd. Run "
                "'python -m pip install -r requirements.txt'."
            ) from exc
        except Exception as exc:
            raise ClarizenPlanError(
                "Could not read the Clarizen .xls file. Confirm it is a valid "
                "Excel export."
            ) from exc
    else:
        try:
            workbook = load_workbook(
                BytesIO(workbook_bytes), data_only=True, read_only=True
            )
        except Exception as exc:
            raise ClarizenPlanError(
                "Could not read the Clarizen Excel file. Confirm it is a valid "
                ".xlsx export."
            ) from exc
        worksheet = (
            workbook["Work Plan"]
            if "Work Plan" in workbook.sheetnames
            else workbook[workbook.sheetnames[0]]
        )
        rows = list(worksheet.iter_rows(values_only=True))

        def convert_date(value):
            return parse_excel_date(value, workbook.epoch)

    header_row, header_map, end_header = find_header_row(rows)
    owner_available = "Owner" in header_map
    resource_available = "Resource" in header_map
    deliverables = []
    warnings = []
    invoicing_section_level = None
    excluded_invoicing_rows = 0

    for row_number, row in enumerate(rows[header_row:], start=header_row + 1):
        def value(column: str):
            index = header_map.get(column)
            return row[index] if index is not None and index < len(row) else None

        state = clean(value("State"))
        name = clean(value("Name"))
        level = parse_level(value("Level"))

        if invoicing_section_level is not None:
            if level is not None and level <= invoicing_section_level:
                invoicing_section_level = None
            else:
                excluded_invoicing_rows += 1
                continue

        if "project invoicing" in name.casefold():
            invoicing_section_level = level
            excluded_invoicing_rows += 1
            continue

        if INVOICING_MILESTONE_RE.match(name):
            excluded_invoicing_rows += 1
            continue

        if not state or state.casefold() != "active":
            continue
        if level is None or level <= 1 or not name:
            continue

        start_date = convert_date(value("Start Date"))
        end_date = convert_date(value(end_header))
        resource = clean(value("Resource")) if resource_available else ""
        fallback_owner = clean(value("Owner")) if owner_available else ""
        owner = resource or fallback_owner
        prefix = f"Excel row {row_number} - {name}"
        if start_date is None:
            warnings.append(f"{prefix}: missing start date.")
        if end_date is None:
            warnings.append(f"{prefix}: missing due/end date.")
        if (resource_available or owner_available) and not owner:
            warnings.append(f"{prefix}: active task is missing an owner/resource.")
        if end_date is not None and end_date < today:
            warnings.append(
                f"{prefix}: active task has an end date in the past "
                f"({end_date.isoformat()})."
            )

        deliverables.append(
            {
                "name": name,
                "owner": owner,
                "state": "Active",
                "start_date": start_date.isoformat() if start_date else "",
                "end_date": end_date.isoformat() if end_date else "",
                "level": level,
                "source": "Clarizen Project Plan",
            }
        )

    if excluded_invoicing_rows:
        warnings.append(
            f"Excluded {excluded_invoicing_rows} Project Invoicing / billing "
            "milestone row(s) from Clarizen deliverables."
        )

    ranked_deliverables = rank_deliverables(
        deliverables,
        reference_date=reference_date,
        recommendation_limit=5,
    )
    return ranked_deliverables, warnings


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Parse a Clarizen project plan Excel export into deliverables JSON."
    )
    parser.add_argument("--input", type=Path, required=True)
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("sample_inputs/clarizen_deliverables.json"),
    )
    parser.add_argument("--limit", type=int, default=5)
    return parser.parse_args()


def main() -> int:
    arguments = parse_args()
    try:
        if arguments.limit < 1:
            raise ClarizenPlanError("--limit must be at least 1.")
        deliverables, warnings = parse_clarizen_workbook(
            arguments.input.read_bytes(),
            reference_date=date.today(),
            file_name=arguments.input.name,
        )
        output = {
            "deliverables": deliverables[: arguments.limit],
            "pm_review_warnings": warnings,
        }
        arguments.output.parent.mkdir(parents=True, exist_ok=True)
        arguments.output.write_text(
            json.dumps(output, indent=2) + "\n", encoding="utf-8"
        )
    except (ClarizenPlanError, OSError) as error:
        print(f"Error: {error}", file=sys.stderr)
        return 2

    print(
        f"Created {arguments.output} with "
        f"{len(output['deliverables'])} deliverable(s)."
    )
    for warning in warnings:
        print(f"- {warning}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
